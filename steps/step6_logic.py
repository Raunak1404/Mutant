"""
step6_logic.py — Step 6: Package & Export

Combine SAP (Step 4) and ESJC (Step 5) outputs into final styled Excel files.
Applies yellow highlighting to split SAP rows and removes internal metadata columns.

Standard interface:
    main(input_path, output_path, libraries_dir) -> dict

Special: The ESJC output is passed via libraries_dir as 'esjc_output.xlsx'.
"""

import json
import sys
import zipfile
from io import BytesIO
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


# Yellow fill for highlighted rows
YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# Internal columns to exclude from SAP output
SAP_EXCLUDE_COLUMNS = ["Action_Count", "Needs_Highlight"]


def main(input_path: str, output_path: str, libraries_dir: str) -> dict:
    """
    Package & Export both SAP and ESJC outputs.

    Args:
        input_path: Path to Step 4 output (SAP formatted data)
        output_path: Path to write final output (ZIP containing both Excel files)
        libraries_dir: Path containing 'esjc_output.xlsx' (Step 5 output)

    Returns:
        dict with keys: success (bool), changelog (list[dict]), stats (dict)
    """
    input_path = Path(input_path)
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    # 1. Load SAP data from Step 4
    sap_df = pd.read_excel(str(input_path))
    sap_rows = len(sap_df)

    # 2. Load ESJC data from libraries_dir
    esjc_path = Path(libraries_dir) / "esjc_output.xlsx"
    if not esjc_path.exists():
        return {
            "success": False,
            "changelog": [],
            "stats": {"error": f"ESJC output not found at {esjc_path}"},
        }

    esjc_df = pd.read_excel(str(esjc_path))
    esjc_rows = len(esjc_df)

    # 3. Identify highlight rows in SAP data
    highlight_mask = None
    if "Needs_Highlight" in sap_df.columns:
        highlight_mask = sap_df["Needs_Highlight"].astype(bool)
        highlight_count = highlight_mask.sum()
    else:
        highlight_count = 0

    # 4. Remove internal columns from SAP data
    sap_visible_cols = [c for c in sap_df.columns if c not in SAP_EXCLUDE_COLUMNS]
    sap_export_df = sap_df[sap_visible_cols].copy()

    # 5. Write SAP Excel with yellow highlighting
    sap_output_path = output_path.parent / "output_sap.xlsx"
    sap_export_df.to_excel(str(sap_output_path), index=False, sheet_name="SAP_Cleaned_Data")

    if highlight_mask is not None and highlight_count > 0:
        wb = load_workbook(str(sap_output_path))
        ws = wb.active
        for row_idx in range(2, len(sap_export_df) + 2):  # +2 for header + 1-indexing
            # row_idx - 2 gives the DataFrame index (0-based)
            df_idx = row_idx - 2
            if df_idx < len(highlight_mask) and highlight_mask.iloc[df_idx]:
                for col_idx in range(1, len(sap_visible_cols) + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = YELLOW_FILL
        wb.save(str(sap_output_path))

    # 6. Write ESJC Excel
    esjc_output_path = output_path.parent / "output_esjc.xlsx"
    esjc_df.to_excel(str(esjc_output_path), index=False, sheet_name="ESJC_Cleaned_Data")

    # 7. Package both into a ZIP file as the primary output
    zip_buffer = BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.write(str(sap_output_path), "SQ_SAP_Vetting_cleanup.xlsx")
        zf.write(str(esjc_output_path), "SQ_ESJC_Vetting_cleanup.xlsx")

    # Write ZIP to output_path (change extension)
    zip_output_path = output_path.with_suffix(".zip")
    zip_output_path.write_bytes(zip_buffer.getvalue())

    # Also write the ZIP to the standard output_path for the orchestrator
    output_path.write_bytes(zip_buffer.getvalue())

    # Clean up intermediate files
    sap_output_path.unlink(missing_ok=True)
    esjc_output_path.unlink(missing_ok=True)

    changelog = [
        {
            "type": "FILE_EXPORTED",
            "column": "SAP",
            "old_value": f"{sap_rows} rows",
            "new_value": f"SQ_SAP_Vetting_cleanup.xlsx ({len(sap_visible_cols)} columns)",
            "reason": f"Exported SAP data with {highlight_count} highlighted rows",
        },
        {
            "type": "FILE_EXPORTED",
            "column": "ESJC",
            "old_value": f"{esjc_rows} rows",
            "new_value": f"SQ_ESJC_Vetting_cleanup.xlsx ({len(esjc_df.columns)} columns)",
            "reason": "Exported ESJC data",
        },
        {
            "type": "ZIP_CREATED",
            "column": "output",
            "old_value": "2 separate files",
            "new_value": "1 ZIP archive",
            "reason": "Packaged both outputs into downloadable ZIP",
        },
    ]

    stats = {
        "sap_rows": sap_rows,
        "sap_columns_visible": len(sap_visible_cols),
        "sap_highlighted_rows": int(highlight_count),
        "esjc_rows": esjc_rows,
        "esjc_columns": len(esjc_df.columns),
        "total_changelog_entries": len(changelog),
    }

    return {
        "success": True,
        "changelog": changelog,
        "stats": stats,
    }


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: python step6_logic.py <input_path> <output_path> [libraries_dir]")
        sys.exit(1)

    inp = sys.argv[1]
    out = sys.argv[2]
    libs = sys.argv[3] if len(sys.argv) > 3 else "./libraries"

    result = main(inp, out, libs)
    print(json.dumps(result.get("stats", {}), indent=2))
    if not result["success"]:
        sys.exit(1)
